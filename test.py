from selenium import webdriver
from time import sleep
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import openpyxl
import time

def setUp():
    global driver
    global wait

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_experimental_option('detach', True)
    driver = webdriver.Chrome( options=chrome_options)
    wait = WebDriverWait(driver, 25)
    driver.maximize_window()

def clickDropdownBoxByText(dropdownBox, textData):
    index =0
    while index < len(dropdownBox):
        if dropdownBox[index].text == textData:
            dropdownBox[index].click()
            break
        index+=1
    if index == len(dropdownBox):
        raise('Error at Dropdown Box! Do not have any data match to dropdown box!')

def fillDataAddBlackCustomer(customerName, phoneNumber, email, customerType, CIF, humanID, dateOfBirth, customerGroup, permanentAddress, contactAddress, role, notation):
    try:
        #Set customerName
        driver.find_element(By.CSS_SELECTOR ,'#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_NAME').send_keys(customerName)
        #Set phoneNumber
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_PHONE').send_keys(phoneNumber)
        #Set email
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_EMAIL').send_keys(email)
        #Set customerType
        dropdownBox = driver.find_elements(By.TAG_NAME, 'option')
        clickDropdownBoxByText(dropdownBox, customerType)
        #Set CIF
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_CIF').send_keys(CIF)
        #Set humanID
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_ID').send_keys(humanID)
        #Set dateOfBirth
        driver.find_element(By.CSS_SELECTOR, '#datetimepicker-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_BIRTH').send_keys(dateOfBirth)
        #Set customerGroup
        dropdownBox = driver.find_elements(By.TAG_NAME, 'option')
        clickDropdownBoxByText(dropdownBox, customerGroup)
        #Set permanentAddress
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_PERMANENT_ADDRESS').send_keys(permanentAddress)
        #Set contactAddress
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_CURRENT_ADDRESS').send_keys(contactAddress)
        #Set role
        driver.find_element(By.CSS_SELECTOR, '#text-input-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_TITLE_NAME').send_keys(role)
        #Set notation
        driver.find_element(By.CSS_SELECTOR, '#textarea-textarea-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:CUS_NOTE').send_keys(notation) 
        #Click 'Lưu' button
        driver.find_element(By.CSS_SELECTOR, '#button-button-LOS_Search_Table_Result_Customer\:LOS_POPUP_RESTRICTED_CUST1\:SAVE').click()
        #check result add success by checking display
        isDisplay = driver.find_element(By.CSS_SELECTOR, '#div_67 > div') 
        if isDisplay.get_attribute('style') == 'display: block;':
            raise Exception('ADD BLACK CUSTOMER TEST CASE FAIL AT Fill data to add black customer FAIL!')
    except:
        raise Exception("[!] ADD BLACK CUSTOMER TEST CASE FAIL AT Fill data to add black customer FAIL!")

def clickQuanlykhachhangdenButton():
    try:
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#launch-menu-item-icon-service > div > div.menu-link-body > a')))    
        listElement = driver.find_elements(By.CLASS_NAME, 'ng-binding');
        for element in listElement:
            if element.get_attribute('title')== 'Quản lý khách hàng đen':
                element.click()
                break;
    except:
        raise Exception("[!] ADD BLACK CUSTOMER TEST CASE FAIL AT CLICK QUAN LY KHACH HANG!")
    

def clickThemkhachhangButtonButton():
    try: 
        iframe = None
        # Wait for iframe loading
        iframe = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'iframe[title="Coach"]')))
        if iframe is None:
            raise Exception("Exception at finding a iframe to add black customer")
        # Switch to iframe which includes 'Thêm khách hàng' button
        driver.switch_to.frame(iframe)
        # Click to 'Thêm 'Thêm khách hàng' button 
        driver.find_element(By.CSS_SELECTOR, '#button-button-LOS_Search_Customer\:Button1').click()
    except:
        raise Exception("[!] ADD BLACK CUSTOMER TEST CASE FAIL AT CLICK QUAN LY KHACH HANG!")        


def testAddBlackCusTomer(customerName, phoneNumber, email, customerType, CIF, humanID, dateOfBirth, 
                         customerGroup, permanentAddress, contactAddress, role, notation):

    clickQuanlykhachhangdenButton()
    clickThemkhachhangButtonButton()
    fillDataAddBlackCustomer(customerName, phoneNumber, email, customerType, CIF, humanID, dateOfBirth, 
                             customerGroup, permanentAddress, contactAddress, role, notation)    


def testLogin( userName, password):
    checkResultLogin = True
    try:
        #Set UserName
        username_input = driver.find_element(By.NAME, 'j_username')  
        username_input.send_keys(userName)
        #Set Password
        password_input = driver.find_element(By.NAME, 'j_password')  
        password_input.send_keys(password)
        #click button Login
        driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div/div[1]/form/a/span").click() 
        driver.find_element(By.ID, 'login-error')
        checkResultLogin = False
    except:
        wait.until(EC.url_changes(driver.current_url))
    return checkResultLogin


def getValueExel(row, col):
    return  sheet.cell(row, col).value


def writeTestCaseResult(testCaseRow, testCaseCol, massage):
    sheet.cell(row=testCaseRow+1, column=testCaseCol, value=str(massage))
    workbook.save("ExelFile/AutoTesting.xlsx")


def executeTestCase(userName, password,customerName, phoneNumber, email, customerType, CIF, humanID, dateOfBirth, 
                    customerGroup, permanentAddress, contactAddress, role, notation, sttTestCase):
    try: 
        driver.get('https://10.0.18.122:9443/ProcessPortal/login.jsp')
        resultTest = testLogin(userName,password)
        if resultTest == False:
            raise Exception(str(sttTestCase) + "[!] LOGIN TEST CASE FA IL!")
        testAddBlackCusTomer(customerName, phoneNumber, email, customerType, CIF, humanID, dateOfBirth, 
                                            customerGroup, permanentAddress, contactAddress, role, notation)
        message = str(sttTestCase) + "[+] TEST CASE SUCCESSFULLY!"
    except Exception as ex:
        message = ex
    finally:
        writeTestCaseResult(sttTestCase, 16, message)


# MAIN FLOW
if __name__ == '__main__':
    #Read date from exel file
    workbook = openpyxl.load_workbook("ExelFile/AutoTesting.xlsx")
    sheet = workbook['LoginTest']
    #Properties of testcase
    totalsCol = sheet.max_column
    #The number of testcase, from 2 to ...
    totalsRow = sheet.max_row
    sttTestCase = 0

    setUp()
    for row in range(2, totalsRow+1):
        sttTestCase = row - 1
        executeTestCase(getValueExel(row,2), getValueExel(row,3), getValueExel(row,4), getValueExel(row,5), 
                        getValueExel(row,6), getValueExel(row,7), getValueExel(row,8), getValueExel(row,9), 
                        getValueExel(row,10), getValueExel(row,11), getValueExel(row,12), getValueExel(row,13), 
                        getValueExel(row,14), getValueExel(row,15), sttTestCase)